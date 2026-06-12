#!/usr/bin/env python3
"""
xlsx2json.py
Converts data/conceptual_graph.xlsx → data/graph.json.

Skips conversion if graph.json is already newer than the XLSX.
Run manually:
    python3 xlsx2json.py
Run forced (ignore mtime):
    python3 xlsx2json.py --force

The JSON schema mirrors exactly what parseCombinedWorkbook() returns in
js/parse/workbook.js so app.js can consume it without SheetJS.
"""

import json
import os
import sys
import re
from datetime import datetime, timezone

XLSX_PATH = "data/conceptual_graph.xlsx"
JSON_PATH = "data/graph.json"


# ── helpers ──────────────────────────────────────────────────────────────────

def normalize_header(h):
    """Mirror normalizeHeader() in js/utils.js: trim, strip BOM, lowercase,
    drop whitespace and underscores."""
    s = str(h or "").strip()
    if s.startswith("﻿"):
        s = s[1:]
    s = s.lower()
    s = re.sub(r"\s+", "", s)
    return s.replace("_", "")


def find_sheet(wb, name):
    """Find a worksheet by name, comparing normalized names (mirrors
    getSheetInfoByName in js/parse/xlsx.js), so "Relation Types" and
    "RelationTypes" both resolve to the same sheet."""
    target = normalize_header(name)
    for sheetname in wb.sheetnames:
        if normalize_header(sheetname) == target:
            return wb[sheetname]
    return None


def normalize_id(val):
    """Mirror normalizeConceptId() in js/utils.js: trim + uppercase."""
    if val is None:
        return ""
    return str(val).strip().upper()


_TRUTHY_RE = re.compile(r"^true|1$", re.IGNORECASE)


def is_truthy(val):
    """Mirror the /^true|1$/i.test(...) checks workbook.js uses for
    isDefault / isDefaultConcept / isDefaultNarrative flags."""
    return bool(_TRUTHY_RE.search(str(val or "")))


def to_number(raw):
    """Mirror `Number(raw) || 0`, collapsing integral floats to ints."""
    if not raw:
        return 0
    try:
        f = float(raw)
    except ValueError:
        return 0
    return int(f) if f.is_integer() else f


def cell(row, *keys):
    """Return the first non-empty value for any of the given keys."""
    for k in keys:
        v = row.get(k)
        if v is not None:
            s = str(v).strip()
            if s:
                return s
    return ""


def parse_key_value_params(raw):
    """Parse a 'key=value; key2=value2' string into a dict (mirrors parseKeyValueParams in workbook.js)."""
    if not raw:
        return {}
    out = {}
    for part in re.split(r"[;\n]", raw):
        part = part.strip()
        if not part:
            continue
        eq = part.find("=")
        if eq > 0:
            k = part[:eq].strip()
            v = part[eq + 1:].strip()
            if k:
                out[k] = v
    return out


def format_cell(v):
    """Stringify a cell value. openpyxl returns numeric cells as float
    (e.g. 1 -> 1.0); collapse integral floats so IDs match the raw string
    JS reads from the XLSX XML (e.g. "1", not "1.0")."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def read_sheet(ws):
    """Convert an openpyxl worksheet to (headers, rows): rows is a list of
    dicts keyed by header."""
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
    rows = []
    for raw_row in raw_rows[1:]:
        d = {}
        for h, v in zip(headers, raw_row):
            if h:
                d[h] = format_cell(v)
        rows.append(d)
    return headers, rows


def sheet_to_rows(ws):
    """Convert an openpyxl worksheet to a list of dicts (header row as keys)."""
    return read_sheet(ws)[1]


def has_columns(headers, names):
    """Mirror hasColumns() in js/parse/xlsx.js."""
    cols = set(normalize_header(h) for h in headers if h)
    return all(normalize_header(n) in cols for n in names)


# ── sheet parsers (mirrors js/parse/workbook.js) ─────────────────────────────

def parse_site(wb):
    ws = find_sheet(wb, "Site")
    if not ws:
        return {}
    out = {}
    for row in sheet_to_rows(ws):
        k = cell(row, "Key", "key")
        v = cell(row, "Value", "value")
        if k:
            out[k] = v
    return out


def parse_concepts_relations(wb):
    """Build outputRows exactly as parseCombinedWorkbook does."""
    concepts_ws = find_sheet(wb, "Concepts")
    relations_ws = find_sheet(wb, "Relations")
    rel_types_ws = find_sheet(wb, "Relation Types")

    if not concepts_ws or not relations_ws:
        return []

    concept_rows = sheet_to_rows(concepts_ws)
    relation_rows = sheet_to_rows(relations_ws)
    rel_type_rows = sheet_to_rows(rel_types_ws) if rel_types_ws else []

    # Build relation types lookup
    relation_types_by_id = {}
    for row in rel_type_rows:
        rid = normalize_id(cell(row, "relationTypeID", "relationID", "RelationID", "id"))
        rname = cell(row, "relationType", "relationName", "RelationType", "name", "label")
        if rid and rname:
            relation_types_by_id[rid] = {
                "relationID": rid,
                "relationType": rname,
                "category": cell(row, "category", "Category", "categoria"),
                "description": cell(row, "description", "Description", "descrição", "descricao"),
            }

    relation_has_type_id = bool(relation_types_by_id)

    # Build concepts lookup
    concepts_by_id = {}
    concept_order = []
    for row in concept_rows:
        cid = normalize_id(cell(row, "conceptID", "ConceptID", "ConceptId", "id"))
        label = cell(row, "ConceptLabel", "conceptLabel", "label", "Concept", "concept")
        if not cid or not label:
            continue
        concepts_by_id[cid] = {
            "id": cid,
            "label": label,
            "description": cell(row, "description", "Description", "descrição", "descricao"),
            "sourceUrl": cell(row, "sourceUrl", "souceUrl", "url", "postUrl"),
            "imagePath": cell(row, "imagePath", "ImagePath", "snapshot", "screenshot"),
            "sourceTitle": cell(row, "sourceTitle", "SourceTitle", "postTitle", "source"),
            "level": cell(row, "level", "Level"),
            "camada": cell(row, "camada", "Camada"),
            "externalRef": cell(row, "externalRef", "ExternalRef", "external_ref", "externalURL", "externalUrl"),
        }
        concept_order.append(cid)

    # Build output rows
    output_rows = []
    concepts_with_outgoing = set()

    for row in relation_rows:
        source_id = normalize_id(cell(row, "ConceptId", "ConceptID", "conceptID", "conceptId", "source"))
        target_id = normalize_id(cell(row, "relatedConcept", "RelatedConcept", "relatedConceptID", "targetConceptId", "target"))
        source = concepts_by_id.get(source_id)
        target = concepts_by_id.get(target_id)
        if not source_id and not target_id:
            continue
        if not source:
            print(f"  ⚠ Relação ignorada: source '{source_id}' sem correspondência em Concepts")
            continue

        relation_type_id = normalize_id(cell(row, "relationTypeID", "RelationTypeID", "relationTypeId", "RelationTypeId", "relationID", "RelationID"))
        relation_type_info = relation_types_by_id.get(relation_type_id) if relation_type_id else None
        relation_name = (
            (relation_type_info["relationType"] if relation_type_info else relation_type_id)
            if relation_has_type_id
            else cell(row, "relationName", "RelationName", "relação", "relacao")
        )

        concepts_with_outgoing.add(source_id)
        output_rows.append({
            "conceptID": source["id"],
            "ConceptLabel": source["label"],
            "concept": source["label"],
            "description": source["description"],
            "relatedConceptID": target_id,
            "relatedConcept": target["label"] if target else target_id,
            "relationTypeID": relation_type_id,
            "relationCategory": relation_type_info["category"] if relation_type_info else "",
            "relationTypeDescription": relation_type_info["description"] if relation_type_info else "",
            "relationName": relation_name or relation_type_id,
            "explanation": cell(row, "explanation", "Explanation", "explicação", "explicacao"),
            "sourceUrl": source["sourceUrl"],
            "imagePath": source["imagePath"],
            "sourceTitle": source["sourceTitle"],
            "level": source["level"],
            "camada": source["camada"],
            "externalRef": source["externalRef"],
        })

    # Concepts with no outgoing relations get a stub row
    for cid in concept_order:
        if cid in concepts_with_outgoing:
            continue
        c = concepts_by_id[cid]
        output_rows.append({
            "conceptID": c["id"], "ConceptLabel": c["label"], "concept": c["label"],
            "description": c["description"], "relatedConceptID": "", "relatedConcept": "",
            "relationTypeID": "", "relationCategory": "", "relationTypeDescription": "",
            "relationName": "", "explanation": "",
            "sourceUrl": c["sourceUrl"], "imagePath": c["imagePath"],
            "sourceTitle": c["sourceTitle"], "level": c["level"],
            "camada": c["camada"], "externalRef": c["externalRef"],
        })

    return output_rows


def parse_narratives(wb):
    """Build narratives store: { byId, order, elementsById, loadedAt }"""
    narratives_ws = find_sheet(wb, "Narratives")
    elements_ws = find_sheet(wb, "Elements")
    if not narratives_ws or not elements_ws:
        return None

    elements_by_id = {}
    for row in sheet_to_rows(elements_ws):
        eid = cell(row, "elementID", "ElementID", "elementId", "id")
        if not eid:
            continue
        elements_by_id[eid] = {
            "elementID": eid,
            "elementTitle": cell(row, "elementTitle", "ElementTitle", "elementName", "title", "titulo", "título"),
            "elementContent": cell(row, "elementContent", "content", "text", "texto"),
            "referencedConceptIDs": cell(row, "referencedConceptIDs", "conceptIDs", "references", "refs", "conceitos"),
        }

    by_id = {}
    order = []
    for row in sheet_to_rows(narratives_ws):
        nid = cell(row, "narrativeID", "NarrativeID", "narrativeId", "id")
        if not nid:
            continue
        seq_raw = cell(row, "elements", "elementIDs", "sequence", "sequencia")
        sequence = [x.strip() for x in seq_raw.split(",") if x.strip()]
        skin_col = cell(row, "skin", "Skin", "skinID")
        hidden_val = cell(row, "hidden", "Hidden", "oculto", "Oculto").lower()
        by_id[nid] = {
            "narrativeID": nid,
            "narrativeTitle": cell(row, "narrativeTitle", "title", "titulo", "título"),
            "narrativeStart": cell(row, "narrativeStart", "start", "startElement") or (sequence[0] if sequence else ""),
            "narrativeSummary": cell(row, "narrativeSummary", "summary", "resumo"),
            "subtitle": cell(row, "subtitle", "Subtitle", "subtítulo"),
            "eyebrow": cell(row, "eyebrow", "Eyebrow", "tag"),
            "year": cell(row, "year", "Year", "ano"),
            "outroQuote": cell(row, "outroQuote", "outro", "closingQuote"),
            "outroMeta": cell(row, "outroMeta", "outroByline", "closingMeta"),
            "elements": sequence,
            "skin": skin_col,
            "hidden": hidden_val in ("1", "true", "sim", "yes"),
        }
        order.append(nid)

    return {
        "byId": by_id,
        "order": order,
        "elementsById": elements_by_id,
        "loadedAt": datetime.now(timezone.utc).isoformat(),
    }


def parse_templates(wb):
    ws = find_sheet(wb, "Templates")
    if not ws:
        return {}
    headers, rows = read_sheet(ws)
    if not has_columns(headers, ["templateID"]):
        return {}
    out = {}
    for row in rows:
        tid = cell(row, "templateID", "TemplateID", "id")
        if not tid:
            continue
        applies_raw = cell(row, "appliesTo", "AppliesTo", "applies")
        out[tid] = {
            "templateID": tid,
            "templateName": cell(row, "templateName", "name", "label"),
            "appliesTo": [s.strip().lower() for s in applies_raw.split(",") if s.strip()] if applies_raw else [],
            "isDefaultConcept": is_truthy(cell(row, "isDefaultConcept", "defaultConcept")),
            "isDefaultNarrative": is_truthy(cell(row, "isDefaultNarrative", "defaultNarrative")),
            "parameters": parse_key_value_params(cell(row, "parameters", "params")),
            "description": cell(row, "description", "Description"),
        }
    return out


def parse_media(wb):
    ws = find_sheet(wb, "Media")
    if not ws:
        return {}
    headers, rows = read_sheet(ws)
    if not has_columns(headers, ["scope", "scopeID"]):
        return {}
    out = {}
    for row in rows:
        scope = cell(row, "scope", "Scope", "kind").lower()
        scope_id = normalize_id(cell(row, "scopeID", "ScopeID", "scopeId", "id"))
        if not scope or not scope_id:
            continue
        key = scope + ":" + scope_id
        pov_raw = cell(row, "povScope", "pov", "povs")
        out.setdefault(key, []).append({
            "order":       to_number(cell(row, "order", "Order", "index", "ord")),
            "type":        (cell(row, "type", "Type", "mediaType") or "image").lower(),
            "file":        cell(row, "file", "File", "filename", "filepath"),
            "poster":      cell(row, "poster", "Poster", "thumbnail", "thumb"),
            "aspectRatio": cell(row, "aspectRatio", "aspect_ratio", "aspect", "ratio"),
            "sandbox":     cell(row, "sandbox", "Sandbox"),
            "caption":     cell(row, "caption", "Caption", "legenda"),
            "sourceUrl":   cell(row, "sourceUrl", "url", "postUrl"),
            "sourceTitle": cell(row, "sourceTitle", "postTitle", "source title", "title"),
            "alt":         cell(row, "alt", "altText", "alternative"),
            "povScope":    [s.strip() for s in pov_raw.split(",") if s.strip()] if pov_raw else [],
        })
    return out


def parse_narrative_skins(wb):
    ws = find_sheet(wb, "Narrative Skins")
    if not ws:
        return {}
    headers, rows = read_sheet(ws)
    if not has_columns(headers, ["skinID", "narrativeID"]):
        return {}
    out = {}
    for row in rows:
        sid = cell(row, "skinID", "SkinID", "id")
        nid = cell(row, "narrativeID", "NarrativeID")
        if not sid or not nid:
            continue
        tags_raw = cell(row, "tags", "Tags")
        eg_mode = cell(row, "egMode", "eg_mode", "EGMode", "egmode")
        out.setdefault(nid, []).append({
            "skinID": sid,
            "narrativeID": nid,
            "skinName": cell(row, "skinName", "name", "label"),
            "isDefault": is_truthy(cell(row, "isDefault", "default")),
            "templateID": cell(row, "templateID", "TemplateID"),
            "parameters": parse_key_value_params(cell(row, "parameters", "params")),
            "coverImage": cell(row, "coverImage", "cover"),
            "egMode": eg_mode or None,
            "tags": [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else [],
        })
    return out


def parse_concept_skins(wb):
    ws = find_sheet(wb, "Concept Skins")
    if not ws:
        return {}
    headers, rows = read_sheet(ws)
    if not has_columns(headers, ["skinID", "conceptID"]):
        return {}
    out = {}
    for row in rows:
        sid = cell(row, "skinID", "SkinID", "id")
        cid = normalize_id(cell(row, "conceptID", "ConceptID"))
        if not sid or not cid:
            continue
        tags_raw = cell(row, "tags", "Tags")
        eg_mode = cell(row, "egMode", "eg_mode", "EGMode")
        out.setdefault(cid, []).append({
            "skinID":         sid,
            "conceptID":      cid,
            "skinName":       cell(row, "skinName", "name", "label") or None,
            "skinImplID":     cell(row, "skinImplID", "implID", "impl") or None,
            "isDefault":      is_truthy(cell(row, "isDefault", "default")),
            "templateID":     cell(row, "templateID", "TemplateID"),
            "parameters":     parse_key_value_params(cell(row, "parameters", "params")),
            "dataSourceType": cell(row, "dataSourceType", "sourceType"),
            "dataSourceID":   cell(row, "dataSourceID", "sourceID"),
            "egMode":         eg_mode or None,
            "tags":           [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else [],
        })
    return out


def parse_concept_texts(wb):
    ws = find_sheet(wb, "ConceptTexts")
    if not ws:
        return {}
    out = {}
    for row in sheet_to_rows(ws):
        cid = normalize_id(cell(row, "conceptID"))
        if not cid:
            continue
        entry = {
            "conceptID": cid,
            "pov": cell(row, "pov"),
            "author": cell(row, "author"),
            "style": cell(row, "style"),
            "lang": cell(row, "lang"),
            "textVersion": cell(row, "textVersion"),
            "isDefault": is_truthy(cell(row, "isDefault", "default")),
            "text": cell(row, "text"),
            "mediaScope": cell(row, "mediaScope"),
        }
        out.setdefault(cid, []).append(entry)
    return out


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    force = "--force" in sys.argv

    if not os.path.exists(XLSX_PATH):
        print(f"✗ {XLSX_PATH} not found.")
        sys.exit(1)

    xlsx_mtime = os.path.getmtime(XLSX_PATH)

    # Skip if JSON is already up to date
    if not force and os.path.exists(JSON_PATH):
        json_mtime = os.path.getmtime(JSON_PATH)
        if json_mtime >= xlsx_mtime:
            print(f"✓ {JSON_PATH} is up to date (skipping).")
            return

    try:
        import openpyxl
    except ImportError:
        print("✗ openpyxl not found. Install with: pip3 install openpyxl")
        sys.exit(1)

    print(f"→ Converting {XLSX_PATH} → {JSON_PATH} …")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    xlsx_mtime_iso = datetime.fromtimestamp(xlsx_mtime, tz=timezone.utc).isoformat()

    rows = parse_concepts_relations(wb)
    narratives = parse_narratives(wb)
    site_config = parse_site(wb)
    templates = parse_templates(wb)
    media = parse_media(wb)
    narrative_skins = parse_narrative_skins(wb)
    concept_skins = parse_concept_skins(wb)
    concept_texts = parse_concept_texts(wb)

    graph_json = {
        "_generated": datetime.now(timezone.utc).isoformat(),
        "_xlsxMtime": xlsx_mtime_iso,
        "rows": rows,
        "narratives": narratives,
        "siteConfig": site_config,
        "media": media,
        "templates": templates,
        "narrativeSkins": narrative_skins,
        "conceptSkins": concept_skins,
        "conceptTexts": concept_texts,
        # skinData is keyed by each site's skins/index.json data contracts
        # (site-owned, fetched at runtime); the JSON fast path leaves it
        # empty and app.js falls back to parseSkinDataContracts() against
        # the XLSX when a skin declares a non-builtin contract.
        "skinData": {},
        "hasGraph": len(rows) > 0,
        "hasNarratives": narratives is not None and len(narratives.get("order", [])) > 0,
    }

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(graph_json, f, ensure_ascii=False, indent=2)

    concepts = len({r["conceptID"] for r in rows if r.get("conceptID")})
    narratives_count = len(narratives["order"]) if narratives else 0
    elements_count = len(narratives["elementsById"]) if narratives else 0
    print(f"✓ Done: {concepts} concepts, {narratives_count} narratives, {elements_count} elements → {JSON_PATH}")


if __name__ == "__main__":
    main()
